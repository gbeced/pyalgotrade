# PyAlgoTrade
#
# Copyright 2011-2015 Gabriel Martin Becedillas Ruiz
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#   http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
.. moduleauthor:: Massimo Fierro <massimo.fierro@gmail.com>
"""

from openpyxl import Workbook
from openpyxl.worksheet import Worksheet
from openpyxl.comments.comments import Comment
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, GradientFill, Border, Side
from openpyxl.styles import Alignment, Protection, Font

from pyalgotrade.stratanalyzer.extendedtrades import ExtendedTradesAnalyzer


class PerformanceReport(object):
    def __init__(self):
        pass

    def writeReport(self, filename, trades):
        wb = Workbook()

        names = wb.sheetnames
        for name in names:
            wb.remove_sheet(wb.get_sheet_by_name(name))

        summarySheet = wb.create_sheet(title="Summary")
        trades_sheet = wb.create_sheet(title="Trades")

        # ----- Trades sheeet -----
        numFormat = "[BLACK][>=0]#,##0.0000;[RED][<0]\\(#,##0.0000\\);General"
        perFormat = "[BLACK][>=0]#0.00%;[RED][<0]\\(#0.00%\);General"

        headerFont = Font(name="Arial", bold=True)
        headerAlign = Alignment(horizontal='center')
        header_fill = PatternFill(
            start_color='AAAAAA', end_color='AAAAAA', fill_type='solid')

        highlightFill = PatternFill(
            start_color='EEEE99', end_color='EEEE99', fill_type='solid')
        highlightBorder = Border(bottom=Side(
            border_style="thin", color="000000"))

        standardFont = Font(name="Arial", size="10")

        for col in range(1, 10):
            trades_sheet.cell(row=1, column=col).font = headerFont
            trades_sheet.cell(row=1, column=col).fill = header_fill
            trades_sheet.cell(row=1, column=col).alignment = headerAlign

        trades_sheet['A1'] = "Trade #\nType"
        trades_sheet['B1'] = "Date"
        trades_sheet['C1'] = "Time"
        trades_sheet['D1'] = "Price"
        trades_sheet['E1'] = "Contracts\nProfit"
        trades_sheet['F1'] = "% Profit\nCum Profit"
        trades_sheet['G1'] = "Run-up\nDrawdown"
        trades_sheet['H1'] = "Entry Eff.\nExit Eff."
        trades_sheet['I1'] = "Total\nEfficiency"

        allTrades = trades.getAll()
        allReturns = trades.getAllReturns()
        enteredOn = trades.allEnterDates
        exitedOn = trades.allExitDates
        longFlags = trades.allLongFlags
        entryPrices = trades.allEntryPrices
        exitPrices = trades.allExitPrices
        allContracts = trades.allContracts
        allCommissions = trades.getCommissionsForAllTrades()

        excelRow = 2
        cumulativeProfit = 0
        cumulativePnL = 0
        cumulativeLosses = 0

        for i in range(0, trades.getCount()):
            for col in range(1, 10):
                trades_sheet.cell(row=excelRow, column=col).font = standardFont
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).font = standardFont

            trades_sheet.cell(row=excelRow, column=1, value=i + 1)
            trades_sheet.cell(row=excelRow, column=1).alignment = Alignment(
                horizontal='center')
            if longFlags[i]:
                buySell = "Buy"
            else:
                buySell = "Sell"
            trades_sheet.cell(row=excelRow + 1, column=1, value=buySell)
            trades_sheet.cell(
                row=excelRow + 1, column=1).alignment = Alignment(
                    horizontal='center')

            entryDate = enteredOn[i]
            exitDate = exitedOn[i]
            trades_sheet.cell(row=excelRow, column=2,
                              value=entryDate.strftime("%Y-%m-%d"))
            trades_sheet.cell(row=excelRow + 1, column=2,
                              value=exitDate.strftime("%Y-%m-%d"))

            trades_sheet.cell(row=excelRow, column=3,
                              value=entryDate.strftime("%H:%M"))
            trades_sheet.cell(row=excelRow + 1, column=3,
                              value=exitDate.strftime("%H:%M"))

            trades_sheet.cell(row=excelRow, column=4, value=entryPrices[i])
            trades_sheet.cell(row=excelRow + 1, column=4, value=exitPrices[i])

            trades_sheet.cell(row=excelRow, column=5, value=allContracts[i])
            trades_sheet.cell(row=excelRow + 1, column=5, value=allReturns[i])
            # TODO(max): Should this include or exclude commissions?
            trades_sheet.cell(
                row=excelRow + 1, column=5).number_format = numFormat

            if longFlags[i]:
                # TODO(max): Check formula with commissions!
                profitPerc = ((
                    entryPrices[i] - allCommissions[i] / allContracts[i]) /
                    exitPrices[i] - 1)
            else:
                # TODO(max): Check formula with commissions!
                profitPerc = - \
                    entryPrices[i] / (exitPrices[i] +
                                      allCommissions[i] / allContracts[i]) - 1
            trades_sheet.cell(row=excelRow, column=6, value=profitPerc)
            trades_sheet.cell(
                row=excelRow, column=6).number_format = perFormat

            # if longFlags[i]:
            #     profit = (exitPrices[i]-entryPrices[i])*allContracts[i]
            # else:
            #     profit = -(exitPrices[i]-entryPrices[i])*allContracts[i]
            cumulativePnL = cumulativeProfit + allReturns[i]
            if allReturns[i] > 0:
                cumulativeProfit = cumulativeProfit + allReturns[i]
            else:
                cumulativeLosses = cumulativeLosses + allReturns[i]
            trades_sheet.cell(row=excelRow + 1, column=6, value=cumulativePnL)
            trades_sheet.cell(
                row=excelRow + 1, column=6).number_format = numFormat

            # Set standard font, and highlight style for 2nd row of trade
            for col in range(1, 10):
                # 1st row
                trades_sheet.cell(row=excelRow, column=col).font = standardFont

                # 2nd row
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).font = standardFont
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).fill = highlightFill
                trades_sheet.cell(row=excelRow + 1,
                                  column=col).border = highlightBorder

            excelRow = excelRow + 2

        # ----- Summary sheeet -----
        titleFont = Font(name="Arial",  size=18, bold=True)
        titleAlign = Alignment(horizontal='center')

        headerFont = Font(name="Arial",  size=14, bold=True)
        headerAlign = Alignment(horizontal='left')

        standardFont = Font(name="Arial", size=10)

        summarySheet['A1'] = "Strategy Performance Report"
        summarySheet.merge_cells("A1:I1")
        summarySheet['A1'].font = titleFont
        summarySheet['A1'].alignment = titleAlign

        summarySheet["B6"] = "Performance Summary: All Trades"
        summarySheet["B6"].font = headerFont
        summarySheet["B6"].alignment = headerAlign

        summarySheet["B8"] = "Net Profits"
        summarySheet["D8"] = cumulativeProfit
        summarySheet["D8"].number_format = numFormat

        summarySheet["F8"] = "Open position P/L"
        summarySheet["H8"] = ""

        summarySheet["B9"] = "Gross Profits"
        summarySheet["D9"] = cumulativeProfit - cumulativeLosses
        summarySheet["D9"].number_format = numFormat
        summarySheet["D9"].comment = Comment(
            "Net profits - Gross losses, i.e. Net profits + Abs(Gross losses)",
            "Report")

        summarySheet["F9"] = "Gross Losses"
        summarySheet["H9"] = cumulativeLosses
        summarySheet["H9"].number_format = numFormat

        summarySheet["B11"] = "Total num. of trades"
        summarySheet["D11"] = trades.getCount()

        summarySheet["F11"] = "Percent profitable"
        if trades.getCount() > 0:
            summarySheet["H11"] = float(
                trades.getProfitableCount()) / float(
                    trades.getCount())
        else:
            summarySheet["H11"] = 0
        summarySheet["H11"].number_format = perFormat

        summarySheet["B12"] = "Num. of winning trades"
        summarySheet["D12"] = trades.getProfitableCount()

        summarySheet["F12"] = "Num. of losing trades"
        summarySheet["H12"] = trades.getUnprofitableCount()

        summarySheet["B14"] = "Largest winning trade"
        if trades.getProfitableCount() > 0:
            summarySheet["D14"] = trades.getPositiveReturns().max()
        else:
            summarySheet["D14"] = 0
        summarySheet["D14"].number_format = numFormat

        summarySheet["F14"] = "Largest losing trade"
        if trades.getUnprofitableCount() > 0:
            summarySheet["H14"] = trades.getNegativeReturns().max()
        else:
            summarySheet["H14"] = 0
        summarySheet["H14"].number_format = numFormat

        if trades.getCount() > 0:
            avgWin = trades.getPositiveReturns().mean()
        else:
            avgWin = 0
        summarySheet["B15"] = "Average winning trade"
        summarySheet["D15"] = avgWin
        summarySheet["D15"].number_format = numFormat

        if trades.getCount() > 0:
            avgLoss = trades.getNegativeReturns().mean()
        else:
            avgLoss = 0
        summarySheet["F15"] = "Average losing trade"
        summarySheet["H15"] = avgLoss
        summarySheet["H15"].number_format = numFormat

        summarySheet["B16"] = "Ratio avg. win/avg. loss"
        if avgLoss > 0:
            summarySheet["D16"] = avgWin / -avgLoss
        else:
            summarySheet["D16"] = 'NaN'
        summarySheet["D16"].number_format = perFormat

        summarySheet["F16"] = "Avg trade (win & loss)"
        if trades.getCount() > 0:
            summarySheet["H16"] = allReturns.mean()
        else:
            summarySheet["H16"] = 0
        summarySheet["H16"].number_format = numFormat

        summarySheet["B21"] = "Max intraday drawdown"
        summarySheet["D21"] = ""

        summarySheet["B22"] = "Profit factor"
        if cumulativeLosses > 0:
            summarySheet["D22"] = (
                cumulativeProfit + cumulativeLosses) / -cumulativeLosses
        else:
            summarySheet["D22"] = 0
        summarySheet["D22"].number_format = perFormat
        summarySheet["D22"].comment = Comment(
            "Gross profits / - Gross losses", "Report")

        summarySheet["F22"] = "Max contracts held"
        summarySheet["H22"] = ""

        summarySheet["B23"] = "Account size required"
        summarySheet["D23"] = ""  # ABS(max intraday drawdown)

        summarySheet["F23"] = "Return on account"
        summarySheet["H23"] = ""  # net profit / account size required

        # Save the file
        wb.save(filename)
